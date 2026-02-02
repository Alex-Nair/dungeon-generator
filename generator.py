import json
import math
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

def generate():
    with open("settings.json", "r") as file:
        settings = json.load(file)

    # Setup - Generating the arrays.
    tiles = [["UNKNOWN" for _ in range(settings["size"])] for _ in range(settings["size"])]
    availableTiles = []

    for i in range(len(tiles)):
        for j in range(len(tiles)):
            availableTiles.append([i, j])

    # Set up the initial start and boss tiles.
    tiles[-1][0] = "START"
    tiles[0][-1] = "BOSS"

    availableTiles.remove([len(tiles) - 1, 0])
    availableTiles.remove([0, len(tiles) - 1])

    if settings["middle_elite_line"]: # Middle line - Across the diagonal going from top-left to bottom-right. Make that diagonal elites.
        for i in range(len(tiles)):
            tiles[i][i] = "ELITE"

            availableTiles.remove([i, i])
    
    # Calculate the counts from the densities.
    counts = {
        "EMPTY": 0,
        "FIGHT": 0,
        "EVENT": 0,
        "ELITE": 0,
        "TREASURE": 0
    }

    # By taking the sum of the densities and divding each one by that sum, we can normalise them such that they all add to one.
    total = 0

    for density in settings["densities"].values():
        total += density
    
    tileTotal = 0
    for densityKey in settings["densities"].keys():
        amount = math.floor(len(availableTiles) * (settings["densities"][densityKey] / total))

        counts[densityKey.upper()] = amount
        tileTotal += amount
    
    # Add extra blanks to fill in any missing space.
    counts["EMPTY"] += len(availableTiles) - tileTotal

    # Now go through each tile type and put them randomly across the grid.
    for tileType in counts.keys():
        for _ in  range(counts[tileType]):
            chosenTile = random.choice(availableTiles)
            tiles[chosenTile[0]][chosenTile[1]] = tileType
            availableTiles.remove(chosenTile)

    # Writing to the workbook.
    workbook = Workbook()
    sheet = workbook.active

    for i in range(len(tiles)):
        for j in range(len(tiles)):
            cell = sheet.cell(row = i + 1, column = j + 1)

            cell.value = tiles[i][j]
            cell.alignment = Alignment(horizontal = "center", vertical = "center")

            colour = {
                "START": "00FF00",
                "EMPTY": "FFFFFF",
                "FIGHT": "FF0000",
                "EVENT": "0000FF",
                "ELITE": "9E00FF",
                "TREASURE": "FFFF00",
                "BOSS": "980000"
            }[tiles[i][j]]

            cell.fill = PatternFill(start_color = colour, end_color = colour, fill_type = "solid")
    
    # Changing cell sizes.
    for column in sheet.columns:
        columnLetter = column[0].column_letter
        sheet.column_dimensions[columnLetter].width = 15
    
    for i in range(len(tiles)):
        sheet.row_dimensions[i + 1].height = 20
    
    workbook.save("map.xlsx")

if __name__ == "__main__":
    generate()